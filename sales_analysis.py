import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

# قراءة البيانات
df = pd.read_excel("sales_data.xlsx", engine="openpyxl")

df["Revenue"] = df["Quantity"] * df["Unit_Price"]
df["Profit"] = df["Revenue"] - (df["Quantity"] * df["Cost"])

product_summary = df.groupby("Product")[["Revenue", "Profit"]].sum()
category_summary = df.groupby("Category")[["Revenue", "Profit"]].sum()

# ======================
# Generate Insights
# ======================

top_product = product_summary["Profit"].idxmax()
top_product_profit = product_summary["Profit"].max()

low_product = product_summary["Profit"].idxmin()
low_product_profit = product_summary["Profit"].min()

top_category = category_summary["Profit"].idxmax()
top_category_profit = category_summary["Profit"].max()

profit_margin = round((df["Profit"].sum() / df["Revenue"].sum()) * 100, 2)

insights = [
    ["Top Product", top_product, top_product_profit],
    ["Lowest Product", low_product, low_product_profit],
    ["Top Category", top_category, top_category_profit],
    ["Overall Profit Margin (%)", profit_margin, ""],
    ["Recommendation",
     f"Focus more on {top_product} and {top_category}. Consider improving or discounting {low_product}.",
     ""]
]

insights_df = pd.DataFrame(
    insights,
    columns=["Metric", "Value", "Amount"]
)

# تصدير البيانات أولًا
with pd.ExcelWriter("sales_report.xlsx", engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Clean_Data", index=False)
    product_summary.to_excel(writer, sheet_name="By_Product")
    category_summary.to_excel(writer, sheet_name="By_Category")
    insights_df.to_excel(writer, sheet_name="Insights", index=False)

# تحميل الملف لإضافة Charts
wb = load_workbook("sales_report.xlsx")

# ======================
# Chart 1: Revenue by Product
# ======================
ws_product = wb["By_Product"]

chart1 = BarChart()
chart1.title = "Revenue by Product"
chart1.y_axis.title = "Revenue"
chart1.x_axis.title = "Product"

data = Reference(
    ws_product,
    min_col=2,
    min_row=1,
    max_row=ws_product.max_row
)
categories = Reference(
    ws_product,
    min_col=1,
    min_row=2,
    max_row=ws_product.max_row
)

chart1.add_data(data, titles_from_data=True)
chart1.set_categories(categories)

ws_product.add_chart(chart1, "E2")

# ======================
# Chart 2: Profit by Category
# ======================
ws_category = wb["By_Category"]

chart2 = BarChart()
chart2.title = "Profit by Category"
chart2.y_axis.title = "Profit"
chart2.x_axis.title = "Category"

data2 = Reference(
    ws_category,
    min_col=2,
    min_row=1,
    max_row=ws_category.max_row
)
categories2 = Reference(
    ws_category,
    min_col=1,
    min_row=2,
    max_row=ws_category.max_row
)

chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(categories2)

ws_category.add_chart(chart2, "E2")

# حفظ الملف
wb.save("sales_report.xlsx")

print("📊 Charts added successfully to sales_report.xlsx")
